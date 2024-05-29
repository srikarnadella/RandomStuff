import random
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

PwordDescription = input("What is the password for: ")
Chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz123456789!@#$%&*"

output = ""

for i in range(35):
    random_integer = random.randint(0, len(Chars))
    output = output + Chars[random_integer]

print(output)


def add_password_to_spreadsheet(filename, description, password):
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Passwords"
        sheet.append(["Description", "Password"])

    sheet = workbook.active

    sheet.append([description, password])

    workbook.save(filename)

filename = "H:\CodingProjects\OneofProjects\PasswordGenerator\passwords.xlsx" 

add_password_to_spreadsheet(filename, PwordDescription, output)

print(f"Password for '{PwordDescription}' added to {filename}")